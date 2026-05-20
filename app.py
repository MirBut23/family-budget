import os
import io
from datetime import datetime
from functools import wraps

from flask import Flask, request, jsonify, session, send_from_directory, send_file
from flask_cors import CORS
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app, supports_credentials=True)

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'familybudget_secret_2025')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Получаем строку подключения из переменной окружения (Render сам подставит)
DATABASE_URL = os.environ.get('DATABASE_URL')

if not DATABASE_URL:
    print("⚠️  ВНИМАНИЕ: Переменная окружения DATABASE_URL не установлена!")
    print("⚠️  Использую локальную БД (только для разработки)")
    DATABASE_URL = 'postgresql://postgres:Mirea23@localhost:5432/familybudget'

def get_db_connection():
    """Подключение к PostgreSQL с поддержкой SSL (нужно для Render)"""
    try:
        # Для Render обязательно нужен sslmode=require
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor, sslmode='require')
        return conn
    except psycopg2.OperationalError as e:
        # Если не получилось с sslmode, пробуем без него (для локальной разработки)
        print(f"Ошибка подключения с sslmode=require: {e}")
        print("Пробую подключиться без sslmode...")
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
        return conn

def init_db():
    """Создание таблиц, если их нет"""
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            email VARCHAR(100) UNIQUE NOT NULL,
            password_hash VARCHAR(200) NOT NULL,
            family_id INTEGER,
            role_label VARCHAR(100),
            is_creator BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cur.execute('''
        CREATE TABLE IF NOT EXISTS families (
            id SERIAL PRIMARY KEY,
            name VARCHAR(100) NOT NULL,
            invite_code VARCHAR(10) UNIQUE NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    cur.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            family_id INTEGER NOT NULL REFERENCES families(id),
            title VARCHAR(200) NOT NULL,
            amount DECIMAL(10,2) NOT NULL,
            type VARCHAR(10) NOT NULL CHECK (type IN ('income', 'expense')),
            category VARCHAR(50),
            date DATE NOT NULL,
            is_hidden BOOLEAN DEFAULT FALSE,
            masked BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Индексы для ускорения запросов
    cur.execute('CREATE INDEX IF NOT EXISTS idx_transactions_family_date ON transactions(family_id, date)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_transactions_user ON transactions(user_id)')
    
    conn.commit()
    cur.close()
    conn.close()
    print("✅ База данных инициализирована (таблицы созданы)")

def generate_invite_code():
    import random
    import string
    return ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))

def hash_password(password):
    import hashlib
    return hashlib.sha256(password.encode()).hexdigest()

def check_password(password, hashed):
    return hash_password(password) == hashed

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Не авторизован'}), 401
        return f(*args, **kwargs)
    return decorated

# ------------------- ОТДАЧА HTML -------------------
@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

# ------------------- API -------------------
@app.route('/api/me', methods=['GET'])
def me():
    if 'user_id' not in session:
        return jsonify({'error': 'not logged in'}), 401
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT u.id, u.name, u.email, u.family_id, u.role_label, u.is_creator,
               f.id as family_id2, f.name as family_name, f.invite_code
        FROM users u
        LEFT JOIN families f ON u.family_id = f.id
        WHERE u.id = %s
    ''', (session['user_id'],))
    user = cur.fetchone()
    cur.close()
    conn.close()
    if user:
        return jsonify({
            'id': user['id'],
            'name': user['name'],
            'email': user['email'],
            'family_id': user['family_id'],
            'role_label': user['role_label'],
            'is_creator': user['is_creator'],
            'family': {
                'id': user['family_id2'],
                'name': user['family_name'],
                'invite_code': user['invite_code']
            } if user['family_id'] else None
        })
    return jsonify({'error': 'User not found'}), 404

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '')
    if not name or not email or not password:
        return jsonify({'error': 'Все поля обязательны'}), 400
    if len(password) < 6:
        return jsonify({'error': 'Пароль должен быть минимум 6 символов'}), 400
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute('INSERT INTO users (name, email, password_hash) VALUES (%s, %s, %s) RETURNING id, name, email',
                    (name, email, hash_password(password)))
        user = cur.fetchone()
        conn.commit()
        session['user_id'] = user['id']
        return jsonify({'user': user})
    except psycopg2.IntegrityError:
        conn.rollback()
        return jsonify({'error': 'Email уже зарегистрирован'}), 400
    finally:
        cur.close()
        conn.close()

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    email = data.get('email', '').strip()
    password = data.get('password', '')
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT id, name, email, password_hash, family_id, role_label, is_creator FROM users WHERE email = %s', (email,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    if not user or not check_password(password, user['password_hash']):
        return jsonify({'error': 'Неверный email или пароль'}), 401
    session['user_id'] = user['id']
    return jsonify({'user': user})

@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/family/create', methods=['POST'])
@login_required
def create_family():
    data = request.json
    family_name = data.get('name', '').strip()
    role_label = data.get('role_label', '')
    if not family_name:
        return jsonify({'error': 'Название семьи обязательно'}), 400
    invite_code = generate_invite_code()
    conn = get_db_connection()
    cur = conn.cursor()
    try:
        cur.execute('INSERT INTO families (name, invite_code) VALUES (%s, %s) RETURNING id, name, invite_code',
                    (family_name, invite_code))
        family = cur.fetchone()
        cur.execute('UPDATE users SET family_id = %s, is_creator = TRUE, role_label = %s WHERE id = %s',
                    (family['id'], role_label, session['user_id']))
        conn.commit()
        return jsonify({'family': family})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route('/api/family/join', methods=['POST'])
@login_required
def join_family():
    data = request.json
    code = data.get('code', '').strip().upper()
    role_label = data.get('role_label', '')
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT id, name FROM families WHERE invite_code = %s', (code,))
    family = cur.fetchone()
    if not family:
        return jsonify({'error': 'Неверный код приглашения'}), 400
    cur.execute('UPDATE users SET family_id = %s, is_creator = FALSE, role_label = %s WHERE id = %s RETURNING id',
                (family['id'], role_label, session['user_id']))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'family': family})

@app.route('/api/family/leave', methods=['POST'])
@login_required
def leave_family():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT is_creator FROM users WHERE id = %s', (session['user_id'],))
    user = cur.fetchone()
    if user and user['is_creator']:
        return jsonify({'error': 'Создатель не может покинуть семью. Сначала передайте права.'}), 400
    cur.execute('UPDATE users SET family_id = NULL, role_label = NULL WHERE id = %s', (session['user_id'],))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/family/members', methods=['GET'])
@login_required
def get_members():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT family_id FROM users WHERE id = %s', (session['user_id'],))
    user = cur.fetchone()
    if not user or not user['family_id']:
        return jsonify([])
    family_id = user['family_id']
    cur.execute('''
        SELECT u.id, u.name, u.role_label, u.is_creator,
               COALESCE(SUM(CASE WHEN t.type = 'income' AND (t.masked = FALSE OR t.user_id = u.id) THEN t.amount ELSE 0 END), 0) as income,
               COALESCE(SUM(CASE WHEN t.type = 'expense' AND (t.masked = FALSE OR t.user_id = u.id) THEN t.amount ELSE 0 END), 0) as expense
        FROM users u
        LEFT JOIN transactions t ON u.id = t.user_id AND t.family_id = %s
        WHERE u.family_id = %s
        GROUP BY u.id, u.name, u.role_label, u.is_creator
        ORDER BY u.is_creator DESC, u.name
    ''', (family_id, family_id))
    members = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify(members)

@app.route('/api/family/member/<int:user_id>/role', methods=['PUT'])
@login_required
def update_role(user_id):
    data = request.json
    role_label = data.get('role_label', '')
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT is_creator, family_id FROM users WHERE id = %s', (session['user_id'],))
    current = cur.fetchone()
    if not current or not current['is_creator']:
        return jsonify({'error': 'Только создатель может менять роли'}), 403
    cur.execute('UPDATE users SET role_label = %s WHERE id = %s AND family_id = %s',
                (role_label, user_id, current['family_id']))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/family/member/<int:user_id>', methods=['DELETE'])
@login_required
def remove_member(user_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT is_creator, family_id FROM users WHERE id = %s', (session['user_id'],))
    current = cur.fetchone()
    if not current or not current['is_creator']:
        return jsonify({'error': 'Только создатель может удалять участников'}), 403
    cur.execute('UPDATE users SET family_id = NULL, role_label = NULL WHERE id = %s AND family_id = %s',
                (user_id, current['family_id']))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/transactions', methods=['GET'])
@login_required
def get_transactions():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT family_id FROM users WHERE id = %s', (session['user_id'],))
    user = cur.fetchone()
    if not user or not user['family_id']:
        return jsonify([])
    family_id = user['family_id']
    member_id = request.args.get('member_id')
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    query = '''
        SELECT t.id, t.title, t.amount, t.type, t.category, t.date, t.is_hidden, t.masked,
               u.id as user_id, u.name as user_name, u.role_label
        FROM transactions t
        JOIN users u ON t.user_id = u.id
        WHERE t.family_id = %s
    '''
    params = [family_id]
    if member_id:
        query += ' AND t.user_id = %s'
        params.append(member_id)
    if from_date:
        query += ' AND t.date >= %s'
        params.append(from_date)
    if to_date:
        query += ' AND t.date <= %s'
        params.append(to_date)
    query += ' ORDER BY t.date DESC, t.id DESC'
    cur.execute(query, params)
    transactions = cur.fetchall()
    cur.close()
    conn.close()
    result = []
    for t in transactions:
        if t['is_hidden'] and t['user_id'] != session['user_id']:
            result.append({
                'id': t['id'],
                'title': 'Личная трата',
                'amount': None,
                'type': 'expense',
                'category': 'Другое',
                'date': str(t['date']),
                'masked': True,
                'user_id': t['user_id'],
                'user_name': t['user_name'],
                'role_label': t['role_label']
            })
        else:
            result.append({
                'id': t['id'],
                'title': t['title'],
                'amount': float(t['amount']),
                'type': t['type'],
                'category': t['category'],
                'date': str(t['date']),
                'is_hidden': t['is_hidden'],
                'masked': False,
                'user_id': t['user_id'],
                'user_name': t['user_name'],
                'role_label': t['role_label']
            })
    return jsonify(result)

@app.route('/api/transactions', methods=['POST'])
@login_required
def add_transaction():
    data = request.json
    title = data.get('title', '').strip()
    amount = data.get('amount')
    tx_type = data.get('type')
    category = data.get('category', 'Другое')
    date_str = data.get('date')
    is_hidden = data.get('is_hidden', False)
    user_id = data.get('user_id', session['user_id'])
    if not title or not amount or not date_str:
        return jsonify({'error': 'Заполните все поля'}), 400
    amount = float(amount)
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT family_id, is_creator FROM users WHERE id = %s', (session['user_id'],))
    current = cur.fetchone()
    if not current or not current['family_id']:
        return jsonify({'error': 'Вы не состоите в семье'}), 400
    family_id = current['family_id']
    if user_id != session['user_id'] and not current['is_creator']:
        return jsonify({'error': 'Только создатель может добавлять операции за других'}), 403
    cur.execute('SELECT id FROM users WHERE id = %s AND family_id = %s', (user_id, family_id))
    if not cur.fetchone():
        return jsonify({'error': 'Пользователь не состоит в вашей семье'}), 400
    cur.execute('''
        INSERT INTO transactions (user_id, family_id, title, amount, type, category, date, is_hidden)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    ''', (user_id, family_id, title, amount, tx_type, category, date_str, is_hidden))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/transactions/<int:tx_id>', methods=['DELETE'])
@login_required
def delete_transaction(tx_id):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('''
        SELECT t.id, t.user_id, u.family_id, u.is_creator
        FROM transactions t
        JOIN users u ON t.user_id = u.id
        WHERE t.id = %s
    ''', (tx_id,))
    tx = cur.fetchone()
    if not tx:
        return jsonify({'error': 'Транзакция не найдена'}), 404
    cur.execute('SELECT family_id, is_creator FROM users WHERE id = %s', (session['user_id'],))
    current = cur.fetchone()
    if tx['family_id'] != current['family_id']:
        return jsonify({'error': 'Нет доступа'}), 403
    if tx['user_id'] != session['user_id'] and not current['is_creator']:
        return jsonify({'error': 'Только создатель или автор могут удалить'}), 403
    cur.execute('DELETE FROM transactions WHERE id = %s', (tx_id,))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/report', methods=['GET'])
@login_required
def get_report():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT family_id FROM users WHERE id = %s', (session['user_id'],))
    user = cur.fetchone()
    if not user or not user['family_id']:
        return jsonify({'error': 'Нет семьи'}), 400
    family_id = user['family_id']
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    member_id = request.args.get('member_id')
    if not from_date or not to_date:
        return jsonify({'error': 'Не указан период'}), 400
    tx_query = '''
        SELECT amount, type, category, user_id, is_hidden
        FROM transactions
        WHERE family_id = %s AND date >= %s AND date <= %s
    '''
    params = [family_id, from_date, to_date]
    if member_id:
        tx_query += ' AND user_id = %s'
        params.append(member_id)
    cur.execute(tx_query, params)
    transactions = cur.fetchall()
    income = 0
    expense = 0
    categories = {}
    members_stats = {}
    for t in transactions:
        if t['is_hidden'] and t['user_id'] != session['user_id']:
            expense += float(t['amount']) if t['type'] == 'expense' else 0
            continue
        amount = float(t['amount'])
        if t['type'] == 'income':
            income += amount
        else:
            expense += amount
            categories[t['category']] = categories.get(t['category'], 0) + amount
        if t['type'] == 'expense':
            members_stats[t['user_id']] = members_stats.get(t['user_id'], 0) + amount
    cur.execute('SELECT id, name FROM users WHERE family_id = %s', (family_id,))
    users = {u['id']: u['name'] for u in cur.fetchall()}
    members_list = [{'id': uid, 'name': users.get(uid, 'Unknown'), 'expense': amt} 
                    for uid, amt in members_stats.items()]
    members_list.sort(key=lambda x: x['expense'], reverse=True)
    categories_list = [{'name': cat, 'amount': amt} for cat, amt in categories.items()]
    categories_list.sort(key=lambda x: x['amount'], reverse=True)
    cur.close()
    conn.close()
    return jsonify({
        'income': income,
        'expense': expense,
        'balance': income - expense,
        'categories': categories_list,
        'members': members_list
    })

# ------------------- ЭКСПОРТ В EXCEL (ТОЛЬКО ДЛЯ СОЗДАТЕЛЯ) -------------------
@app.route('/api/report/excel', methods=['GET'])
@login_required
def get_report_excel():
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute('SELECT family_id, is_creator, name FROM users WHERE id = %s', (session['user_id'],))
    user = cur.fetchone()
    
    if not user or not user['family_id']:
        cur.close()
        conn.close()
        return jsonify({'error': 'Вы не состоите в семье'}), 400
    
    if not user['is_creator']:
        cur.close()
        conn.close()
        return jsonify({'error': 'Только создатель семьи может скачать отчёт в Excel'}), 403
    
    family_id = user['family_id']
    current_user_name = user['name']
    
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    member_id = request.args.get('member_id')
    
    if not from_date or not to_date:
        cur.close()
        conn.close()
        return jsonify({'error': 'Не указан период'}), 400
    
    tx_query = '''
        SELECT t.amount, t.type, t.category, t.date, t.title,
               u.name as user_name, u.role_label, t.is_hidden
        FROM transactions t
        JOIN users u ON t.user_id = u.id
        WHERE t.family_id = %s AND t.date >= %s AND t.date <= %s
    '''
    params = [family_id, from_date, to_date]
    if member_id:
        tx_query += ' AND t.user_id = %s'
        params.append(member_id)
    tx_query += ' ORDER BY t.date DESC'
    
    cur.execute(tx_query, params)
    transactions = cur.fetchall()
    
    cur.execute('SELECT name FROM families WHERE id = %s', (family_id,))
    family = cur.fetchone()
    
    cur.execute('SELECT name, role_label FROM users WHERE family_id = %s', (family_id,))
    members = cur.fetchall()
    
    cur.close()
    conn.close()
    
    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "Транзакции"
    
    headers = ['Дата', 'Участник', 'Роль', 'Тип', 'Категория', 'Название', 'Сумма, ₽', 'Примечание']
    ws1.append(headers)
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2d5a27", end_color="2d5a27", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in range(1, len(headers) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    income_total = 0.0
    expense_total = 0.0
    
    for tx in transactions:
        amount = float(tx['amount'])
        if tx['type'] == 'income':
            income_total += amount
            amount_str = f"+ {amount:,.2f}"
        else:
            expense_total += amount
            amount_str = f"- {amount:,.2f}"
        
        is_masked = tx['is_hidden'] and tx['user_name'] != current_user_name
        note = "🔒 Скрытая трата" if is_masked else ""
        
        ws1.append([
            str(tx['date']),
            tx['user_name'],
            tx['role_label'] or 'Участник',
            'Доход' if tx['type'] == 'income' else 'Расход',
            tx['category'],
            tx['title'] if not is_masked else "Личная трата",
            amount_str,
            note
        ])
    
    ws1.append([])
    ws1.append(['', '', '', '', '', 'ИТОГО ДОХОДЫ:', f"+ {income_total:,.2f}", ''])
    ws1.append(['', '', '', '', '', 'ИТОГО РАСХОДЫ:', f"- {expense_total:,.2f}", ''])
    ws1.append(['', '', '', '', '', 'ОСТАТОК:', f"{income_total - expense_total:,.2f}", ''])
    
    ws2 = wb.create_sheet("По категориям")
    
    cat_query = '''
        SELECT category, SUM(amount) as total
        FROM transactions
        WHERE family_id = %s AND date >= %s AND date <= %s AND type = 'expense'
        GROUP BY category
        ORDER BY total DESC
    '''
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(cat_query, [family_id, from_date, to_date])
    cats = cur.fetchall()
    cur.close()
    conn.close()
    
    ws2.append(['Категория', 'Сумма расходов, ₽', 'Доля от всех расходов, %'])
    
    for col in range(1, 4):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for cat in cats:
        total = float(cat['total']) if cat['total'] else 0
        percent = round(total / expense_total * 100, 1) if expense_total > 0 else 0
        ws2.append([cat['category'], f"{total:,.2f}", f"{percent}%"])
    
    ws3 = wb.create_sheet("По участникам")
    
    member_query = '''
        SELECT u.name, u.role_label,
               SUM(CASE WHEN t.type = 'income' AND (t.is_hidden = FALSE OR t.user_id = u.id) THEN t.amount ELSE 0 END) as income,
               SUM(CASE WHEN t.type = 'expense' AND (t.is_hidden = FALSE OR t.user_id = u.id) THEN t.amount ELSE 0 END) as expense
        FROM users u
        LEFT JOIN transactions t ON u.id = t.user_id AND t.family_id = %s AND t.date >= %s AND t.date <= %s
        WHERE u.family_id = %s
        GROUP BY u.id, u.name, u.role_label
        ORDER BY expense DESC
    '''
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(member_query, [family_id, from_date, to_date, family_id])
    members_stats = cur.fetchall()
    cur.close()
    conn.close()
    
    ws3.append(['Участник', 'Роль', 'Доходы, ₽', 'Расходы, ₽', 'Чистый результат, ₽'])
    
    for col in range(1, 6):
        cell = ws3.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for m in members_stats:
        income = float(m['income']) if m['income'] else 0
        expense = float(m['expense']) if m['expense'] else 0
        balance = income - expense
        ws3.append([
            m['name'],
            m['role_label'] or 'Участник',
            f"{income:,.2f}",
            f"{expense:,.2f}",
            f"{balance:,.2f}"
        ])
    
    for sheet in [ws1, ws2, ws3]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'budget_report_{from_date}_to_{to_date}.xlsx'
    )

# ------------------- ЗАПУСК -------------------
if __name__ == '__main__':
    # Сначала инициализируем базу данных
    init_db()
    
    # Затем запускаем сервер
    port = int(os.environ.get('PORT', 5000))
    print("✅ База данных инициализирована")
    print(f"🌐 Сервер запущен: http://localhost:{port}")
    app.run(debug=False, host='0.0.0.0', port=port)